"""
NYC LL97 Carbon Emissions Calculator
Flask web application
"""

import os
import json
import sqlite3
from flask import Flask, render_template, request, jsonify, session, redirect
from flask_cors import CORS

# Import LL97 data
import sys
sys.path.insert(0, os.path.dirname(__file__))
from data.emission_factors import (
    DEFAULT_UTILITY_FACTORS, ESPM_LIMITS, ESPM_PROPERTY_TYPES,
    OCCUPANCY_GROUP_LIMITS, UNIT_CONVERSIONS, PENALTY_RATE,
    COMPLIANCE_PERIODS, PERIOD_LABELS, MODIFIABLE_PERIODS, MODIFIABLE_FUELS
)
from data.db_setup import get_db_connection, import_ll84_data, DB_PATH
from data.saved_db import (
    get_saved_db_connection, create_saved_tables, SAVED_DB_PATH
)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'll97-calc-secret-key-change-in-prod')
CORS(app)

# ---------------------------------------------------------------------------
# ESPM property type name normalizer
# LL84 uses the same ESPM names as LL97, but spacing/capitalisation can vary.
# Build a lowercase lookup once at startup.
# ---------------------------------------------------------------------------
_ESPM_LOWER = {t.lower(): t for t in ESPM_PROPERTY_TYPES}

# Explicit alias table for known LL84 variations that don't exactly match
_ESPM_ALIASES = {
    'multi-family housing':           'Multifamily Housing',
    'multi family housing':           'Multifamily Housing',
    'k12 school':                     'K-12 School',
    'k-12 school':                    'K-12 School',
    'parking garage':                 'Parking',
    'parking lot':                    'Parking',
    'non-refrigerated warehouse':     'Non-Refrigerated Warehouse',
    'non refrigerated warehouse':     'Non-Refrigerated Warehouse',
    'refrigerated warehouse':         'Refrigerated Warehouse',
    'fitness center':                 'Fitness Center/Health Club/Gym',
    'health club':                    'Fitness Center/Health Club/Gym',
    'gym':                            'Fitness Center/Health Club/Gym',
    'urgent care':                    'Urgent Care/Clinic/Other Outpatient',
    'clinic':                         'Urgent Care/Clinic/Other Outpatient',
    'post office':                    'Mailing Center/Post Office',
    'mailing center':                 'Mailing Center/Post Office',
    'supermarket':                    'Supermarket/Grocery Store',
    'grocery store':                  'Supermarket/Grocery Store',
    'worship':                        'Worship Facility',
    'house of worship':               'Worship Facility',
    'self storage':                   'Self-Storage Facility',
    'self-storage':                   'Self-Storage Facility',
    'data center':                    'Data Center',
    'social hall':                    'Social/Meeting Hall',
    'meeting hall':                   'Social/Meeting Hall',
    'performing arts':                'Performing Arts',
    'theater':                        'Movie Theater',
    'movie theater':                  'Movie Theater',
    'senior care':                    'Senior Care Community',
    'assisted living':                'Senior Care Community',
    'residential care':               'Residential Care Facility',
    'dormitory':                      'Residence Hall/Dormitory',
    'residence hall':                 'Residence Hall/Dormitory',
    'medical office':                 'Medical Office',
    'financial office':               'Financial Office',
    'bank':                           'Bank Branch',
    'bank branch':                    'Bank Branch',
    'convenience store':              'Convenience Store without Gas Station',
    'laboratory':                     'Laboratory',
    'library':                        'Library',
    'museum':                         'Museum',
    'courthouse':                     'Courthouse',
    'pre-school':                     'Pre-school/Daycare',
    'daycare':                        'Pre-school/Daycare',
    'vocational school':              'Vocational School',
    'college':                        'College/University',
    'university':                     'College/University',
    'adult education':                'Adult Education',
    'ambulatory surgical':            'Ambulatory Surgical Center',
    'automobile dealership':          'Automobile Dealership',
    'car dealership':                 'Automobile Dealership',
    'bowling alley':                  'Bowling Alley',
    'distribution center':            'Distribution Center',
    'warehouse':                      'Non-Refrigerated Warehouse',
    'enclosed mall':                  'Enclosed Mall',
    'lifestyle center':               'Lifestyle Center',
    'strip mall':                     'Strip Mall',
    'wholesale club':                 'Wholesale Club/Supercenter',
    'supercenter':                    'Wholesale Club/Supercenter',
    'transportation terminal':        'Transportation Terminal/Station',
    'transit station':                'Transportation Terminal/Station',
    'manufacturing':                  'Manufacturing/Industrial Plant',
    'industrial':                     'Manufacturing/Industrial Plant',
    'food sales':                     'Food Sales',
    'food service':                   'Food Service',
    'restaurant':                     'Restaurant',
    'hospital':                       'Hospital (General Medical & Surgical)',
}

def normalize_espm_type(raw):
    """Map a raw LL84 ESPM property type string to the canonical LL97 name."""
    if not raw:
        return ''
    # Exact match (case-insensitive, strip whitespace)
    key = raw.strip().lower()
    if key in _ESPM_LOWER:
        return _ESPM_LOWER[key]
    # Check alias table
    if key in _ESPM_ALIASES:
        return _ESPM_ALIASES[key]
    # No match — return as-is so the user can see what came from LL84
    return raw.strip()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Default energy prices and annual escalators (% per year, compounding)
DEFAULT_ENERGY_PRICES = {
    'electricity_kwh':    {'label': 'Electricity',      'unit': '$/kWh',   'price': 0.20,  'escalator': 3.0},
    'natural_gas_therm':  {'label': 'Natural Gas',      'unit': '$/therm', 'price': 1.50,  'escalator': 3.0},
    'district_steam_mlb': {'label': 'District Steam',   'unit': '$/mLb',   'price': 18.00, 'escalator': 3.0},
    'fuel_oil_2_gal':     {'label': '#2 Fuel Oil',      'unit': '$/gal',   'price': 3.00,  'escalator': 3.0},
    'fuel_oil_4_gal':     {'label': '#4 Fuel Oil',      'unit': '$/gal',   'price': 2.90,  'escalator': 3.0},
}


def get_energy_prices_config():
    """Return current energy prices + escalators, merging session overrides."""
    user = session.get('energy_prices', {})
    result = {}
    for fuel, defaults in DEFAULT_ENERGY_PRICES.items():
        result[fuel] = {
            'price':     float(user.get(fuel, {}).get('price',     defaults['price'])),
            'escalator': float(user.get(fuel, {}).get('escalator', defaults['escalator'])),
        }
    return result


def get_escalated_price(base_price, escalator_pct, year):
    """Price in given year = base * (1 + pct/100)^(year - 2024)."""
    return base_price * ((1 + escalator_pct / 100.0) ** (year - 2024))


def get_utility_factors(session_settings=None):
    """Return utility factors, merging any user overrides from session."""
    import copy
    factors = copy.deepcopy(DEFAULT_UTILITY_FACTORS)
    if session_settings:
        for period in MODIFIABLE_PERIODS:
            if period in session_settings:
                for fuel, val in session_settings[period].items():
                    try:
                        factors[period][fuel] = float(val)
                    except (ValueError, TypeError):
                        pass
    return factors


def calculate_emissions(energy, utility_factors):
    """
    Calculate annual GHG emissions (tCO2e) for each compliance period.
    energy dict keys:
        electricity_kwh, natural_gas_therms, district_steam_mlb,
        fuel_oil_2_gal, fuel_oil_4_gal
    Returns dict of period -> tCO2e
    """
    uc = UNIT_CONVERSIONS
    # Convert to kBtu for combustion fuels
    ng_kbtu = (energy.get('natural_gas_therms') or 0) * uc['natural_gas_therm_to_kbtu']
    steam_kbtu = (energy.get('district_steam_mlb') or 0) * uc['district_steam_mlb_to_kbtu']
    fo2_kbtu = (energy.get('fuel_oil_2_gal') or 0) * uc['fuel_oil_2_gal_to_kbtu']
    fo4_kbtu = (energy.get('fuel_oil_4_gal') or 0) * uc['fuel_oil_4_gal_to_kbtu']
    elec_kwh = energy.get('electricity_kwh') or 0

    results = {}
    for period in COMPLIANCE_PERIODS:
        f = utility_factors[period]
        emissions = (
            elec_kwh    * f['electricity_kwh'] +
            ng_kbtu     * f['natural_gas_kbtu'] +
            steam_kbtu  * f['district_steam_kbtu'] +
            fo2_kbtu    * f['fuel_oil_2_kbtu'] +
            fo4_kbtu    * f['fuel_oil_4_kbtu']
        )
        # Breakdown by source
        results[period] = {
            'total': round(emissions, 4),
            'breakdown': {
                'electricity':    round(elec_kwh   * f['electricity_kwh'],   4),
                'natural_gas':    round(ng_kbtu    * f['natural_gas_kbtu'],   4),
                'district_steam': round(steam_kbtu * f['district_steam_kbtu'],4),
                'fuel_oil_2':     round(fo2_kbtu   * f['fuel_oil_2_kbtu'],   4),
                'fuel_oil_4':     round(fo4_kbtu   * f['fuel_oil_4_kbtu'],   4),
            }
        }
    return results


def calculate_limit(occupancy_groups, period):
    """
    Calculate building emissions limit for a given period.
    occupancy_groups: list of {'property_type': str, 'floor_area': float}
    Returns tCO2e limit for the period.
    """
    total_limit = 0.0
    for og in occupancy_groups:
        pt = og.get('property_type', '')
        area = float(og.get('floor_area') or 0)
        if pt in ESPM_LIMITS and area > 0:
            factor = ESPM_LIMITS[pt].get(period, 0)
            total_limit += factor * area
    return round(total_limit, 4)


def calculate_penalty(emissions, limit):
    """Calculate annual penalty in dollars."""
    overage = emissions - limit
    if overage <= 0:
        return 0.0
    return round(overage * PENALTY_RATE, 2)


def calculate_utility_costs(energy, prices):
    """
    Calculate annual utility costs.
    prices dict: electricity_kwh, natural_gas_therm, district_steam_mlb,
                 fuel_oil_2_gal, fuel_oil_4_gal
    """
    costs = {
        'electricity':    (energy.get('electricity_kwh') or 0)      * (prices.get('electricity_kwh') or 0),
        'natural_gas':    (energy.get('natural_gas_therms') or 0)   * (prices.get('natural_gas_therm') or 0),
        'district_steam': (energy.get('district_steam_mlb') or 0)   * (prices.get('district_steam_mlb') or 0),
        'fuel_oil_2':     (energy.get('fuel_oil_2_gal') or 0)       * (prices.get('fuel_oil_2_gal') or 0),
        'fuel_oil_4':     (energy.get('fuel_oil_4_gal') or 0)       * (prices.get('fuel_oil_4_gal') or 0),
    }
    costs['total'] = sum(costs.values())
    return {k: round(v, 2) for k, v in costs.items()}


def _compute_building_compliance(energy, occupancy_groups):
    """
    Compute per-period baseline compliance for a building.
    Returns a dict keyed by period with emissions, limit, overage, penalty, compliant.
    """
    user_overrides = session.get('utility_factors', {})
    utility_factors = get_utility_factors(user_overrides)
    emissions_by_period = calculate_emissions(energy, utility_factors)
    total_floor_area = sum(og['floor_area'] for og in occupancy_groups) or 1
    results = {}
    for period in COMPLIANCE_PERIODS:
        emissions = emissions_by_period[period]['total']
        limit = calculate_limit(occupancy_groups, period)
        overage = max(0, emissions - limit)
        penalty = calculate_penalty(emissions, limit)
        results[period] = {
            'label':      PERIOD_LABELS[period],
            'emissions':  round(emissions, 4),
            'limit':      round(limit, 4),
            'overage':    round(overage, 4),
            'penalty':    round(penalty, 2),
            'compliant':  emissions <= limit,
            'intensity_kg':       round(emissions / total_floor_area * 1000, 4),
            'limit_intensity_kg': round(limit    / total_floor_area * 1000, 4),
        }
    return results


def _compute_scenario_period_compliance(energy, occupancy_groups, scenario_id, conn):
    """
    Compute per-period compliance for a building under a given scenario.
    Accumulates measure savings up to the START of each compliance period.
    Returns a dict keyed by period.
    """
    user_overrides = session.get('utility_factors', {})
    utility_factors = get_utility_factors(user_overrides)

    placements = conn.execute('''
        SELECT sm.year, m.elec_savings, m.gas_savings, m.steam_savings,
               m.oil2_savings, m.oil4_savings
        FROM scenario_measures sm
        JOIN measures m ON sm.measure_id = m.id
        WHERE sm.scenario_id = ?
        ORDER BY sm.year
    ''', (scenario_id,)).fetchall()

    # Build cumulative savings by year
    savings_by_year = {}
    for p in placements:
        yr = p['year']
        if yr not in savings_by_year:
            savings_by_year[yr] = dict(elec=0, gas=0, steam=0, oil2=0, oil4=0)
        savings_by_year[yr]['elec']  += (p['elec_savings']  or 0)
        savings_by_year[yr]['gas']   += (p['gas_savings']   or 0)
        savings_by_year[yr]['steam'] += (p['steam_savings'] or 0)
        savings_by_year[yr]['oil2']  += (p['oil2_savings']  or 0)
        savings_by_year[yr]['oil4']  += (p['oil4_savings']  or 0)

    # Period start years
    period_starts = {
        '2024_2029': 2024,
        '2030_2034': 2030,
        '2035_2039': 2035,
        '2040_2049': 2040,
        '2050_plus': 2050,
    }

    total_floor_area = sum(og['floor_area'] for og in occupancy_groups) or 1
    results = {}
    for period in COMPLIANCE_PERIODS:
        start_year = period_starts[period]
        # Accumulate all measure savings up to (but not including) this period's start
        cumulative = dict(elec=0, gas=0, steam=0, oil2=0, oil4=0)
        for yr, savings in savings_by_year.items():
            if yr < start_year:
                for k in cumulative:
                    cumulative[k] += savings[k]
        # Adjust energy by cumulative savings
        adj = {
            'electricity_kwh':    max(0, (energy.get('electricity_kwh')    or 0) - cumulative['elec']),
            'natural_gas_therms': max(0, (energy.get('natural_gas_therms') or 0) - cumulative['gas']),
            'district_steam_mlb': max(0, (energy.get('district_steam_mlb') or 0) - cumulative['steam']),
            'fuel_oil_2_gal':     max(0, (energy.get('fuel_oil_2_gal')     or 0) - cumulative['oil2']),
            'fuel_oil_4_gal':     max(0, (energy.get('fuel_oil_4_gal')     or 0) - cumulative['oil4']),
        }
        emissions_adj = calculate_emissions(adj, utility_factors)[period]['total']
        limit = calculate_limit(occupancy_groups, period)
        overage = max(0, emissions_adj - limit)
        penalty = calculate_penalty(emissions_adj, limit)
        results[period] = {
            'label':      PERIOD_LABELS[period],
            'emissions':  round(emissions_adj, 4),
            'limit':      round(limit, 4),
            'overage':    round(overage, 4),
            'penalty':    round(penalty, 2),
            'compliant':  emissions_adj <= limit,
            'intensity_kg':       round(emissions_adj / total_floor_area * 1000, 4),
            'limit_intensity_kg': round(limit          / total_floor_area * 1000, 4),
        }
    return results


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return redirect('/select-building')


@app.route('/select-building')
def select_building_page():
    return render_template('select_building.html',
        espm_property_types=ESPM_PROPERTY_TYPES,
        period_labels=PERIOD_LABELS,
        compliance_periods=COMPLIANCE_PERIODS,
    )


@app.route('/calculate')
def calculate_page():
    return render_template('calculate_result.html',
        compliance_periods=COMPLIANCE_PERIODS,
        period_labels=PERIOD_LABELS,
    )


@app.route('/portfolio')
def portfolio_page():
    return render_template('portfolio.html')


@app.route('/manage')
def manage():
    return render_template('manage.html',
        compliance_periods=COMPLIANCE_PERIODS,
        period_labels=PERIOD_LABELS,
    )


@app.route('/settings')
def settings():
    user_factors = session.get('utility_factors', {})
    return render_template('settings.html',
        default_factors=DEFAULT_UTILITY_FACTORS,
        user_factors=user_factors,
        modifiable_periods=MODIFIABLE_PERIODS,
        modifiable_fuels=MODIFIABLE_FUELS,
        period_labels=PERIOD_LABELS,
    )


@app.route('/api/settings', methods=['POST'])
def save_settings():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    session['utility_factors'] = data
    return jsonify({'status': 'saved'})


@app.route('/api/settings/reset', methods=['POST'])
def reset_settings():
    session.pop('utility_factors', None)
    return jsonify({'status': 'reset'})


@app.route('/api/settings/current')
def get_settings():
    user_factors = session.get('utility_factors', {})
    factors = get_utility_factors(user_factors)
    return jsonify({
        'current': factors,
        'defaults': DEFAULT_UTILITY_FACTORS,
        'user_overrides': user_factors,
    })


@app.route('/api/settings/prices', methods=['GET'])
def get_price_settings():
    return jsonify({
        'prices': get_energy_prices_config(),
        'defaults': {k: {'price': v['price'], 'escalator': v['escalator']}
                     for k, v in DEFAULT_ENERGY_PRICES.items()},
    })


@app.route('/api/settings/prices', methods=['POST'])
def save_price_settings():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    session['energy_prices'] = data
    return jsonify({'status': 'saved'})


@app.route('/api/settings/prices/reset', methods=['POST'])
def reset_price_settings():
    session.pop('energy_prices', None)
    return jsonify({'status': 'reset'})


# ---------------------------------------------------------------------------
# Saved Buildings list/detail endpoints
# ---------------------------------------------------------------------------

@app.route('/saved-buildings')
def saved_buildings_page():
    return redirect('/portfolio')


@app.route('/api/saved-buildings')
def list_saved_buildings():
    if not os.path.exists(SAVED_DB_PATH):
        return jsonify({'buildings': []})
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    rows = conn.execute(
        'SELECT save_name, property_name, address, borough, postcode, '
        'gross_floor_area, saved_at, selected_scenario_id FROM saved_buildings ORDER BY saved_at DESC'
    ).fetchall()
    conn.close()
    return jsonify({'buildings': [dict(r) for r in rows]})


@app.route('/api/saved-buildings/<path:name>')
def get_saved_building(name):
    if not os.path.exists(SAVED_DB_PATH):
        return jsonify({'error': 'Not found'}), 404
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    row = conn.execute(
        'SELECT * FROM saved_buildings WHERE save_name = ?', (name,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = dict(row)
    try:
        d['occupancy_groups'] = json.loads(d.get('occupancy_groups') or '[]')
    except Exception:
        d['occupancy_groups'] = []
    try:
        d['compliance_cache'] = json.loads(d.get('compliance_cache') or 'null')
    except Exception:
        d['compliance_cache'] = None
    return jsonify({'building': d})


@app.route('/api/buildings/<path:name>/select-scenario', methods=['POST'])
def set_selected_scenario(name):
    """Set (or clear) the selected scenario for a saved building."""
    if not os.path.exists(SAVED_DB_PATH):
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json() or {}
    scenario_id = data.get('scenario_id')  # None/null to unstar

    conn = get_saved_db_connection()
    create_saved_tables(conn)
    row = conn.execute('SELECT * FROM saved_buildings WHERE save_name = ?', (name,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Building not found'}), 404

    conn.execute(
        'UPDATE saved_buildings SET selected_scenario_id = ? WHERE save_name = ?',
        (scenario_id, name)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'selected_scenario_id': scenario_id})


@app.route('/api/portfolio')
def get_portfolio():
    """Return all saved buildings with baseline compliance cache + selected scenario info."""
    if not os.path.exists(SAVED_DB_PATH):
        return jsonify({'buildings': []})
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    rows = conn.execute(
        'SELECT * FROM saved_buildings ORDER BY saved_at DESC'
    ).fetchall()

    buildings = []
    for row in rows:
        b = dict(row)
        try:
            b['occupancy_groups'] = json.loads(b.get('occupancy_groups') or '[]')
        except Exception:
            b['occupancy_groups'] = []
        try:
            b['compliance_cache'] = json.loads(b.get('compliance_cache') or 'null')
        except Exception:
            b['compliance_cache'] = None

        # Compute scenario compliance if a selected scenario exists
        b['selected_scenario_compliance'] = None
        if b.get('selected_scenario_id'):
            try:
                energy = {
                    'electricity_kwh':    b.get('electricity_kwh')    or 0,
                    'natural_gas_therms': b.get('natural_gas_therms') or 0,
                    'district_steam_mlb': b.get('district_steam_mlb') or 0,
                    'fuel_oil_2_gal':     b.get('fuel_oil_2_gal')     or 0,
                    'fuel_oil_4_gal':     b.get('fuel_oil_4_gal')     or 0,
                }
                scen_row = conn.execute(
                    'SELECT name FROM scenarios WHERE id = ?', (b['selected_scenario_id'],)
                ).fetchone()
                b['selected_scenario_name'] = scen_row['name'] if scen_row else None
                b['selected_scenario_compliance'] = _compute_scenario_period_compliance(
                    energy, b['occupancy_groups'], b['selected_scenario_id'], conn
                )
            except Exception:
                pass

        # Fill in missing compliance_cache from stored building data
        if b['compliance_cache'] is None:
            occ = b['occupancy_groups']
            energy = {
                'electricity_kwh':    b.get('electricity_kwh')    or 0,
                'natural_gas_therms': b.get('natural_gas_therms') or 0,
                'district_steam_mlb': b.get('district_steam_mlb') or 0,
                'fuel_oil_2_gal':     b.get('fuel_oil_2_gal')     or 0,
                'fuel_oil_4_gal':     b.get('fuel_oil_4_gal')     or 0,
            }
            if occ and any(v > 0 for v in energy.values()):
                try:
                    cache = _compute_building_compliance(energy, occ)
                    b['compliance_cache'] = cache
                    conn.execute(
                        'UPDATE saved_buildings SET compliance_cache = ? WHERE save_name = ?',
                        (json.dumps(cache), b['save_name'])
                    )
                    conn.commit()
                except Exception:
                    pass

        buildings.append(b)
    conn.close()
    return jsonify({'buildings': buildings})


@app.route('/api/portfolio/recalculate-all', methods=['POST'])
def recalculate_portfolio():
    """Recompute compliance_cache for every building from its stored energy/occupancy data."""
    if not os.path.exists(SAVED_DB_PATH):
        return jsonify({'status': 'ok', 'updated': 0})
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    rows = conn.execute('SELECT * FROM saved_buildings').fetchall()
    updated = 0
    for row in rows:
        b = dict(row)
        try:
            occ = json.loads(b.get('occupancy_groups') or '[]')
        except Exception:
            occ = []
        energy = {
            'electricity_kwh':    b.get('electricity_kwh')    or 0,
            'natural_gas_therms': b.get('natural_gas_therms') or 0,
            'district_steam_mlb': b.get('district_steam_mlb') or 0,
            'fuel_oil_2_gal':     b.get('fuel_oil_2_gal')     or 0,
            'fuel_oil_4_gal':     b.get('fuel_oil_4_gal')     or 0,
        }
        if occ and any(v > 0 for v in energy.values()):
            try:
                cache = _compute_building_compliance(energy, occ)
                conn.execute(
                    'UPDATE saved_buildings SET compliance_cache = ? WHERE save_name = ?',
                    (json.dumps(cache), b['save_name'])
                )
                updated += 1
            except Exception:
                pass
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'updated': updated})


@app.route('/api/buildings/<path:name>/refresh-compliance', methods=['POST'])
def refresh_building_compliance(name):
    """Recompute and save compliance_cache for one building from its stored data."""
    if not os.path.exists(SAVED_DB_PATH):
        return jsonify({'error': 'Not found'}), 404
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    row = conn.execute('SELECT * FROM saved_buildings WHERE save_name = ?', (name,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    b = dict(row)
    try:
        occ = json.loads(b.get('occupancy_groups') or '[]')
    except Exception:
        occ = []
    energy = {
        'electricity_kwh':    b.get('electricity_kwh')    or 0,
        'natural_gas_therms': b.get('natural_gas_therms') or 0,
        'district_steam_mlb': b.get('district_steam_mlb') or 0,
        'fuel_oil_2_gal':     b.get('fuel_oil_2_gal')     or 0,
        'fuel_oil_4_gal':     b.get('fuel_oil_4_gal')     or 0,
    }
    if not occ or not any(v > 0 for v in energy.values()):
        conn.close()
        return jsonify({'error': 'Insufficient data'}), 400
    try:
        cache = _compute_building_compliance(energy, occ)
        conn.execute(
            'UPDATE saved_buildings SET compliance_cache = ? WHERE save_name = ?',
            (json.dumps(cache), name)
        )
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'compliance_cache': cache})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/scenario-period-compliance', methods=['POST'])
def scenario_period_compliance():
    """Compute per-period compliance for a building under a given scenario."""
    data = request.get_json()
    scenario_id = data.get('scenario_id')
    if not scenario_id:
        return jsonify({'error': 'scenario_id required'}), 400

    energy = {
        'electricity_kwh':    _safe_float(data.get('electricity_kwh'))    or 0,
        'natural_gas_therms': _safe_float(data.get('natural_gas_therms')) or 0,
        'district_steam_mlb': _safe_float(data.get('district_steam_mlb')) or 0,
        'fuel_oil_2_gal':     _safe_float(data.get('fuel_oil_2_gal'))     or 0,
        'fuel_oil_4_gal':     _safe_float(data.get('fuel_oil_4_gal'))     or 0,
    }
    occupancy_groups = []
    for og in (data.get('occupancy_groups') or []):
        pt   = og.get('property_type', '').strip()
        area = _safe_float(og.get('floor_area'))
        if pt and area and area > 0:
            occupancy_groups.append({'property_type': pt, 'floor_area': area})
    if not occupancy_groups:
        return jsonify({'error': 'occupancy_groups required'}), 400

    conn = get_saved_db_connection()
    result = _compute_scenario_period_compliance(energy, occupancy_groups, scenario_id, conn)
    # Also fetch scenario name
    scen_row = conn.execute('SELECT name FROM scenarios WHERE id = ?', (scenario_id,)).fetchone()
    conn.close()
    return jsonify({
        'results': result,
        'scenario_name': scen_row['name'] if scen_row else None,
    })


@app.route('/api/search')
def search_buildings():
    """Search LL84 and savedbuildings databases by address, BBL, BIN, or name."""
    q = request.args.get('q', '').strip()
    if len(q) < 3:
        return jsonify({'results': []})

    q_like = f'%{q}%'
    results = []

    # ── 1. Query savedbuildings.db first so saved buildings appear at the top ──
    if os.path.exists(SAVED_DB_PATH):
        try:
            sconn = get_saved_db_connection()
            srows = sconn.execute('''
                SELECT * FROM saved_buildings
                WHERE save_name LIKE ?
                   OR UPPER(address)       LIKE UPPER(?)
                   OR UPPER(property_name) LIKE UPPER(?)
                   OR source_bbl = ?
                   OR source_bin = ?
                ORDER BY saved_at DESC
                LIMIT 10
            ''', (q_like, q_like, q_like, q, q)).fetchall()
            sconn.close()
            for row in srows:
                r = dict(row)
                results.append({
                    'source':             'saved',
                    'save_name':          r['save_name'],
                    'bbl':                r['source_bbl'] or '',
                    'bin':                r['source_bin'] or '',
                    'property_name':      r['property_name'] or '',
                    'address':            r['address'] or '',
                    'borough':            r['borough'] or '',
                    'postcode':           r['postcode'] or '',
                    'year_ending':        r['year_ending'] or '',
                    'gross_floor_area':   r['gross_floor_area'],
                    'energy_star_score':  r['energy_star_score'] or '',
                    'electricity_kwh':    r['electricity_kwh'],
                    'natural_gas_therms': r['natural_gas_therms'],
                    'district_steam_mlb': r['district_steam_mlb'],
                    'fuel_oil_2_gal':     r['fuel_oil_2_gal'],
                    'fuel_oil_4_gal':     r['fuel_oil_4_gal'],
                    'occupancy_groups':   json.loads(r['occupancy_groups'] or '[]'),
                    'reported_ghg':       r['reported_ghg'],
                })
        except Exception:
            pass

    # ── 2. Query LL84 database ─────────────────────────────────────────────────
    if not os.path.exists(DB_PATH):
        if not results:
            return jsonify({'results': [], 'warning': 'LL84 database not yet initialized. Run db_setup.py.'})
        return jsonify({'results': results})

    conn = get_db_connection()
    rows = conn.execute('''
        SELECT bbl, bin, property_name, address, borough, postcode,
               year_ending, gross_floor_area,
               primary_property_type, primary_floor_area,
               second_property_type,  second_floor_area,
               third_property_type,   third_floor_area,
               electricity_kwh, natural_gas_kbtu, district_steam_kbtu,
               fuel_oil_2_kbtu, fuel_oil_4_kbtu,
               reported_ghg_emissions, energy_star_score
        FROM buildings
        WHERE bbl = ?
           OR bin = ?
           OR UPPER(address) LIKE UPPER(?)
           OR UPPER(property_name) LIKE UPPER(?)
        ORDER BY year_ending DESC, property_name
        LIMIT 20
    ''', (q, q, q_like, q_like)).fetchall()
    conn.close()

    for row in rows:
        r = dict(row)

        # Convert stored kBtu values to natural units for the calculator form
        def kbtu_to(val, divisor):
            return round(val / divisor, 1) if val else None

        ng_therms  = kbtu_to(r['natural_gas_kbtu'],   100)      # kBtu → therms
        steam_mlb  = kbtu_to(r['district_steam_kbtu'], 1194)     # kBtu → Mlb
        fo2_gal    = kbtu_to(r['fuel_oil_2_kbtu'],     138.5)    # kBtu → gallons
        fo4_gal    = kbtu_to(r['fuel_oil_4_kbtu'],     146.0)    # kBtu → gallons
        # electricity_kwh is already stored as kWh (converted on import)
        elec_kwh   = round(r['electricity_kwh'], 0) if r['electricity_kwh'] else None

        # Build occupancy groups list (up to 3 uses from LL84).
        total_gfa = r['gross_floor_area'] or 0
        occupancy_groups = []
        for type_col, area_col in [
            ('primary_property_type', 'primary_floor_area'),
            ('second_property_type',  'second_floor_area'),
            ('third_property_type',   'third_floor_area'),
        ]:
            raw_type  = r.get(type_col) or ''
            area      = r.get(area_col)
            norm_type = normalize_espm_type(raw_type)
            if not norm_type:
                continue
            floor_area = round(float(area), 0) if (area is not None and float(area) > 0) else None
            occupancy_groups.append({
                'property_type': norm_type,
                'floor_area':    floor_area,
            })

        # Floor-area fallbacks when individual areas are missing
        if occupancy_groups:
            any_area = any(g['floor_area'] for g in occupancy_groups)
            if not any_area and total_gfa:
                occupancy_groups[0]['floor_area'] = round(total_gfa, 0)
        else:
            if total_gfa:
                occupancy_groups.append({'property_type': '', 'floor_area': round(total_gfa, 0)})

        results.append({
            'source':           'll84',
            'bbl':              r['bbl'],
            'bin':              r['bin'],
            'property_name':    r['property_name'],
            'address':          r['address'],
            'borough':          r['borough'],
            'postcode':         r['postcode'],
            'year_ending':      r['year_ending'],
            'gross_floor_area': total_gfa,
            'electricity_kwh':  elec_kwh,
            'natural_gas_therms': ng_therms,
            'district_steam_mlb': steam_mlb,
            'fuel_oil_2_gal':   fo2_gal,
            'fuel_oil_4_gal':   fo4_gal,
            'occupancy_groups': occupancy_groups,
            'reported_ghg':     r['reported_ghg_emissions'],
            'energy_star_score': r['energy_star_score'],
        })

    return jsonify({'results': results})


@app.route('/api/save-building', methods=['POST'])
def save_building():
    """Save or overwrite a building in savedbuildings.db."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400

    save_name = (data.get('save_name') or '').strip()
    overwrite  = bool(data.get('overwrite', False))
    b          = data.get('building', {})

    if not save_name:
        return jsonify({'error': 'save_name is required'}), 400

    conn = get_saved_db_connection()
    create_saved_tables(conn)

    existing = conn.execute(
        'SELECT id FROM saved_buildings WHERE save_name = ?', (save_name,)
    ).fetchone()

    if existing and not overwrite:
        conn.close()
        return jsonify({'error': 'name_exists'}), 409

    # Compute compliance cache at save time
    occ_groups = b.get('occupancy_groups') or []
    energy_for_cache = {
        'electricity_kwh':    _safe_float(b.get('electricity_kwh'))    or 0,
        'natural_gas_therms': _safe_float(b.get('natural_gas_therms')) or 0,
        'district_steam_mlb': _safe_float(b.get('district_steam_mlb')) or 0,
        'fuel_oil_2_gal':     _safe_float(b.get('fuel_oil_2_gal'))     or 0,
        'fuel_oil_4_gal':     _safe_float(b.get('fuel_oil_4_gal'))     or 0,
    }
    compliance_cache_val = None
    if occ_groups and any(energy_for_cache.values()):
        try:
            compliance_cache_val = json.dumps(
                _compute_building_compliance(energy_for_cache, occ_groups)
            )
        except Exception:
            pass

    row = {
        'save_name':          save_name,
        'source_bbl':         b.get('bbl') or '',
        'source_bin':         b.get('bin') or '',
        'property_name':      b.get('property_name') or '',
        'address':            b.get('address') or '',
        'borough':            b.get('borough') or '',
        'postcode':           b.get('postcode') or '',
        'year_ending':        b.get('year_ending') or '',
        'gross_floor_area':   _safe_float(b.get('gross_floor_area')),
        'energy_star_score':  b.get('energy_star_score') or '',
        'electricity_kwh':    _safe_float(b.get('electricity_kwh')),
        'natural_gas_therms': _safe_float(b.get('natural_gas_therms')),
        'district_steam_mlb': _safe_float(b.get('district_steam_mlb')),
        'fuel_oil_2_gal':     _safe_float(b.get('fuel_oil_2_gal')),
        'fuel_oil_4_gal':     _safe_float(b.get('fuel_oil_4_gal')),
        'occupancy_groups':   json.dumps(occ_groups),
        'reported_ghg':       _safe_float(b.get('reported_ghg')),
        'compliance_cache':   compliance_cache_val,
    }

    if existing and overwrite:
        conn.execute('''
            UPDATE saved_buildings SET
                saved_at = datetime('now'),
                source_bbl = :source_bbl, source_bin = :source_bin,
                property_name = :property_name, address = :address,
                borough = :borough, postcode = :postcode, year_ending = :year_ending,
                gross_floor_area = :gross_floor_area, energy_star_score = :energy_star_score,
                electricity_kwh = :electricity_kwh, natural_gas_therms = :natural_gas_therms,
                district_steam_mlb = :district_steam_mlb, fuel_oil_2_gal = :fuel_oil_2_gal,
                fuel_oil_4_gal = :fuel_oil_4_gal, occupancy_groups = :occupancy_groups,
                reported_ghg = :reported_ghg, compliance_cache = :compliance_cache
            WHERE save_name = :save_name
        ''', row)
    else:
        conn.execute('''
            INSERT INTO saved_buildings (
                save_name, source_bbl, source_bin, property_name, address,
                borough, postcode, year_ending, gross_floor_area, energy_star_score,
                electricity_kwh, natural_gas_therms, district_steam_mlb,
                fuel_oil_2_gal, fuel_oil_4_gal, occupancy_groups, reported_ghg,
                compliance_cache
            ) VALUES (
                :save_name, :source_bbl, :source_bin, :property_name, :address,
                :borough, :postcode, :year_ending, :gross_floor_area, :energy_star_score,
                :electricity_kwh, :natural_gas_therms, :district_steam_mlb,
                :fuel_oil_2_gal, :fuel_oil_4_gal, :occupancy_groups, :reported_ghg,
                :compliance_cache
            )
        ''', row)

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'save_name': save_name})


@app.route('/api/calculate', methods=['POST'])
def calculate():
    """Main calculation endpoint."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    # Energy inputs
    energy = {
        'electricity_kwh':      _safe_float(data.get('electricity_kwh')),
        'natural_gas_therms':   _safe_float(data.get('natural_gas_therms')),
        'district_steam_mlb':   _safe_float(data.get('district_steam_mlb')),
        'fuel_oil_2_gal':       _safe_float(data.get('fuel_oil_2_gal')),
        'fuel_oil_4_gal':       _safe_float(data.get('fuel_oil_4_gal')),
    }

    # Occupancy groups (up to 4)
    occupancy_groups = []
    for og in (data.get('occupancy_groups') or []):
        pt = og.get('property_type', '').strip()
        area = _safe_float(og.get('floor_area'))
        if pt and area and area > 0:
            occupancy_groups.append({'property_type': pt, 'floor_area': area})

    if not occupancy_groups:
        return jsonify({'error': 'At least one occupancy group with floor area is required'}), 400

    total_floor_area = sum(og['floor_area'] for og in occupancy_groups)

    # Energy prices for cost analysis — use session settings (base year 2024 prices)
    prices_cfg = get_energy_prices_config()
    prices = {k: v['price'] for k, v in prices_cfg.items()}

    # Get utility factors (with user overrides from session)
    user_overrides = session.get('utility_factors', {})
    utility_factors = get_utility_factors(user_overrides)

    # Calculate emissions
    emissions_by_period = calculate_emissions(energy, utility_factors)

    # Calculate results for each period
    results = {}
    for period in COMPLIANCE_PERIODS:
        emissions = emissions_by_period[period]['total']
        limit = calculate_limit(occupancy_groups, period)
        overage = max(0, emissions - limit)
        penalty = calculate_penalty(emissions, limit)
        compliant = emissions <= limit

        intensity = round(emissions / total_floor_area * 1000, 4) if total_floor_area else 0  # kgCO2e/sf
        limit_intensity = round(limit / total_floor_area * 1000, 4) if total_floor_area else 0

        results[period] = {
            'label':             PERIOD_LABELS[period],
            'emissions':         emissions,
            'limit':             limit,
            'overage':           round(overage, 4),
            'penalty':           penalty,
            'compliant':         compliant,
            'intensity_kg':      intensity,        # kgCO2e/sf/yr
            'limit_intensity_kg': limit_intensity,
            'breakdown':         emissions_by_period[period]['breakdown'],
        }

    # Utility costs
    utility_costs = calculate_utility_costs(energy, prices)

    return jsonify({
        'results':           results,
        'utility_costs':     utility_costs,
        'total_floor_area':  total_floor_area,
        'energy_summary':    energy,
    })


@app.route('/api/espm-types')
def espm_types():
    return jsonify({'types': ESPM_PROPERTY_TYPES})


# ---------------------------------------------------------------------------
# Reduction Plan routes
# ---------------------------------------------------------------------------

def _get_period_for_year(year):
    """Map a calendar year to its LL97 compliance period key."""
    if year <= 2029: return '2024_2029'
    if year <= 2034: return '2030_2034'
    if year <= 2039: return '2035_2039'
    if year <= 2049: return '2040_2049'
    return '2050_plus'


@app.route('/reduction-plan')
def reduction_plan():
    return render_template('reduction_plan.html')


# ── Measures ────────────────────────────────────────────────────────────────

@app.route('/api/measures')
def get_measures():
    building = request.args.get('building', '').strip()
    if not building:
        return jsonify({'error': 'building param required'}), 400
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    rows = conn.execute(
        'SELECT * FROM measures WHERE building_save_name = ? ORDER BY created_at, id',
        (building,)
    ).fetchall()
    conn.close()
    return jsonify({'measures': [dict(r) for r in rows]})


@app.route('/api/measures', methods=['POST'])
def create_measure():
    data = request.get_json()
    building = (data.get('building_save_name') or '').strip()
    name     = (data.get('name') or '').strip()
    if not building or not name:
        return jsonify({'error': 'building_save_name and name required'}), 400
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    cur = conn.execute('''
        INSERT INTO measures
            (building_save_name, name, cost, elec_savings, gas_savings,
             steam_savings, oil2_savings, oil4_savings)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        building, name,
        _safe_float(data.get('cost'))          or 0,
        _safe_float(data.get('elec_savings'))  or 0,
        _safe_float(data.get('gas_savings'))   or 0,
        _safe_float(data.get('steam_savings')) or 0,
        _safe_float(data.get('oil2_savings'))  or 0,
        _safe_float(data.get('oil4_savings'))  or 0,
    ))
    row = conn.execute('SELECT * FROM measures WHERE id = ?', (cur.lastrowid,)).fetchone()
    conn.commit()
    conn.close()
    return jsonify({'measure': dict(row)}), 201


@app.route('/api/measures/<int:measure_id>', methods=['DELETE'])
def delete_measure(measure_id):
    conn = get_saved_db_connection()
    conn.execute('DELETE FROM scenario_measures WHERE measure_id = ?', (measure_id,))
    conn.execute('DELETE FROM measures WHERE id = ?', (measure_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/measures/<int:measure_id>', methods=['PUT'])
def update_measure(measure_id):
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400
    conn = get_saved_db_connection()
    conn.execute('''
        UPDATE measures SET
            name = ?, cost = ?, elec_savings = ?, gas_savings = ?,
            steam_savings = ?, oil2_savings = ?, oil4_savings = ?
        WHERE id = ?
    ''', (
        name,
        _safe_float(data.get('cost'))          or 0,
        _safe_float(data.get('elec_savings'))  or 0,
        _safe_float(data.get('gas_savings'))   or 0,
        _safe_float(data.get('steam_savings')) or 0,
        _safe_float(data.get('oil2_savings'))  or 0,
        _safe_float(data.get('oil4_savings'))  or 0,
        measure_id,
    ))
    conn.commit()
    row = conn.execute('SELECT * FROM measures WHERE id = ?', (measure_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'measure': dict(row)})


# ── Scenarios ───────────────────────────────────────────────────────────────

@app.route('/api/scenarios')
def get_scenarios():
    building = request.args.get('building', '').strip()
    if not building:
        return jsonify({'error': 'building param required'}), 400
    conn = get_saved_db_connection()
    create_saved_tables(conn)
    rows = conn.execute(
        'SELECT id, name, number, created_at, updated_at FROM scenarios '
        'WHERE building_save_name = ? ORDER BY number',
        (building,)
    ).fetchall()
    # Find out which scenario is selected for this building
    bldg_row = conn.execute(
        'SELECT selected_scenario_id FROM saved_buildings WHERE save_name = ?', (building,)
    ).fetchone()
    selected_id = bldg_row['selected_scenario_id'] if bldg_row else None
    conn.close()
    scenarios = [dict(r) for r in rows]
    for s in scenarios:
        s['is_selected'] = (s['id'] == selected_id)
    return jsonify({'scenarios': scenarios, 'selected_scenario_id': selected_id})


@app.route('/api/scenarios/<int:scenario_id>')
def get_scenario(scenario_id):
    conn = get_saved_db_connection()
    scenario = conn.execute(
        'SELECT * FROM scenarios WHERE id = ?', (scenario_id,)
    ).fetchone()
    if not scenario:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    placements = conn.execute('''
        SELECT sm.id, sm.measure_id, sm.year,
               m.name AS measure_name, m.cost,
               m.elec_savings, m.gas_savings, m.steam_savings,
               m.oil2_savings, m.oil4_savings
        FROM scenario_measures sm
        JOIN measures m ON sm.measure_id = m.id
        WHERE sm.scenario_id = ?
        ORDER BY sm.year, m.name
    ''', (scenario_id,)).fetchall()
    conn.close()
    return jsonify({
        'scenario':   dict(scenario),
        'placements': [dict(p) for p in placements],
    })


@app.route('/api/scenarios/<int:scenario_id>/rename', methods=['POST'])
def rename_scenario(scenario_id):
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400
    conn = get_saved_db_connection()
    conn.execute('UPDATE scenarios SET name = ? WHERE id = ?', (name, scenario_id))
    conn.commit()
    scenario = conn.execute('SELECT * FROM scenarios WHERE id = ?', (scenario_id,)).fetchone()
    conn.close()
    if not scenario:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'scenario': dict(scenario)})


@app.route('/api/scenarios/<int:scenario_id>', methods=['DELETE'])
def delete_scenario(scenario_id):
    conn = get_saved_db_connection()
    scenario = conn.execute('SELECT * FROM scenarios WHERE id = ?', (scenario_id,)).fetchone()
    if not scenario:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    # Clear selected_scenario_id on the building if it points to this scenario
    conn.execute(
        'UPDATE saved_buildings SET selected_scenario_id = NULL WHERE selected_scenario_id = ?',
        (scenario_id,)
    )
    conn.execute('DELETE FROM scenario_measures WHERE scenario_id = ?', (scenario_id,))
    conn.execute('DELETE FROM scenarios WHERE id = ?', (scenario_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'deleted'})


@app.route('/api/scenarios/save', methods=['POST'])
def save_scenario():
    data               = request.get_json()
    building           = (data.get('building_save_name') or '').strip()
    scenario_id        = data.get('scenario_id')   # None → create new
    placements         = data.get('placements', []) # [{measure_id, year}]

    conn = get_saved_db_connection()
    create_saved_tables(conn)

    if scenario_id:
        conn.execute(
            "UPDATE scenarios SET updated_at = datetime('now') WHERE id = ?",
            (scenario_id,)
        )
    else:
        if not building:
            conn.close()
            return jsonify({'error': 'building_save_name required for new scenario'}), 400
        max_num = conn.execute(
            'SELECT COALESCE(MAX(number), 0) FROM scenarios WHERE building_save_name = ?',
            (building,)
        ).fetchone()[0]
        new_num = max_num + 1
        cur = conn.execute(
            'INSERT INTO scenarios (building_save_name, name, number) VALUES (?, ?, ?)',
            (building, f'Scenario {new_num}', new_num)
        )
        scenario_id = cur.lastrowid

    # Replace placements
    conn.execute('DELETE FROM scenario_measures WHERE scenario_id = ?', (scenario_id,))
    for p in placements:
        mid  = p.get('measure_id')
        year = p.get('year')
        if mid and year:
            conn.execute(
                'INSERT INTO scenario_measures (scenario_id, measure_id, year) VALUES (?, ?, ?)',
                (scenario_id, mid, year)
            )

    conn.commit()
    scenario = conn.execute('SELECT * FROM scenarios WHERE id = ?', (scenario_id,)).fetchone()
    conn.close()
    return jsonify({'scenario': dict(scenario)})


# ── Scenario chart computation ───────────────────────────────────────────────

@app.route('/api/scenario-compute', methods=['POST'])
def scenario_compute():
    data        = request.get_json()
    scenario_id = data.get('scenario_id')
    if not scenario_id:
        return jsonify({'error': 'scenario_id required'}), 400

    energy = {
        'electricity_kwh':    _safe_float(data.get('electricity_kwh'))    or 0,
        'natural_gas_therms': _safe_float(data.get('natural_gas_therms')) or 0,
        'district_steam_mlb': _safe_float(data.get('district_steam_mlb')) or 0,
        'fuel_oil_2_gal':     _safe_float(data.get('fuel_oil_2_gal'))     or 0,
        'fuel_oil_4_gal':     _safe_float(data.get('fuel_oil_4_gal'))     or 0,
    }
    occupancy_groups = []
    for og in (data.get('occupancy_groups') or []):
        pt   = og.get('property_type', '').strip()
        area = _safe_float(og.get('floor_area'))
        if pt and area and area > 0:
            occupancy_groups.append({'property_type': pt, 'floor_area': area})
    if not occupancy_groups:
        return jsonify({'error': 'occupancy_groups required'}), 400

    prices_cfg = get_energy_prices_config()  # year-0 (2024) prices + escalators

    conn = get_saved_db_connection()
    placements = conn.execute('''
        SELECT sm.year, m.name, m.cost,
               m.elec_savings, m.gas_savings, m.steam_savings,
               m.oil2_savings, m.oil4_savings
        FROM scenario_measures sm
        JOIN measures m ON sm.measure_id = m.id
        WHERE sm.scenario_id = ?
        ORDER BY sm.year, m.name
    ''', (scenario_id,)).fetchall()
    conn.close()

    # Accumulate savings, measure costs, and measure names per year
    savings_by_year      = {}
    measure_cost_by_year = {}
    measure_names_by_year = {}
    for p in placements:
        yr = p['year']
        if yr not in savings_by_year:
            savings_by_year[yr] = dict(elec=0, gas=0, steam=0, oil2=0, oil4=0)
        measure_names_by_year.setdefault(yr, []).append(p['name'])
        savings_by_year[yr]['elec']  += (p['elec_savings']  or 0)
        savings_by_year[yr]['gas']   += (p['gas_savings']   or 0)
        savings_by_year[yr]['steam'] += (p['steam_savings'] or 0)
        savings_by_year[yr]['oil2']  += (p['oil2_savings']  or 0)
        savings_by_year[yr]['oil4']  += (p['oil4_savings']  or 0)
        measure_cost_by_year[yr] = measure_cost_by_year.get(yr, 0) + (p['cost'] or 0)

    user_overrides  = session.get('utility_factors', {})
    utility_factors = get_utility_factors(user_overrides)
    cumulative      = dict(elec=0, gas=0, steam=0, oil2=0, oil4=0)
    yearly_data     = []

    for year in range(2024, 2051):
        if year in savings_by_year:
            for k in cumulative:
                cumulative[k] += savings_by_year[year][k]

        adj = {
            'electricity_kwh':    max(0, energy['electricity_kwh']    - cumulative['elec']),
            'natural_gas_therms': max(0, energy['natural_gas_therms'] - cumulative['gas']),
            'district_steam_mlb': max(0, energy['district_steam_mlb'] - cumulative['steam']),
            'fuel_oil_2_gal':     max(0, energy['fuel_oil_2_gal']     - cumulative['oil2']),
            'fuel_oil_4_gal':     max(0, energy['fuel_oil_4_gal']     - cumulative['oil4']),
        }
        period       = _get_period_for_year(year)
        emissions    = calculate_emissions(adj, utility_factors)[period]['total']
        limit        = calculate_limit(occupancy_groups, period)
        fine         = calculate_penalty(emissions, limit)
        year_prices  = {k: get_escalated_price(v['price'], v['escalator'], year)
                        for k, v in prices_cfg.items()}
        energy_costs = calculate_utility_costs(adj, year_prices)
        yearly_data.append({
            'year':         year,
            'emissions':    round(emissions, 4),
            'limit':        round(limit, 4),
            'fine':         round(fine, 2),
            'energy_cost':  energy_costs,
            'measure_cost':  round(measure_cost_by_year.get(year, 0), 2),
            'measure_names': measure_names_by_year.get(year, []),
        })

    return jsonify({'yearly_data': yearly_data})


# ── Suggest Plan ──────────────────────────────────────────────────────────────

@app.route('/api/suggest-plan', methods=['POST'])
def suggest_plan():
    data     = request.get_json()
    building = (data.get('building_save_name') or '').strip()
    if not building:
        return jsonify({'error': 'building_save_name required'}), 400

    mode         = data.get('mode', 'minimize_cost')   # 'minimize_cost' | 'minimize_fines'
    scope        = data.get('scope', 'all')            # 'all' | 'unplaced'
    use_discount = bool(data.get('use_discount', False))
    discount_rate = float(data.get('discount_rate') or 5) / 100.0

    energy = {
        'electricity_kwh':    _safe_float(data.get('electricity_kwh'))    or 0,
        'natural_gas_therms': _safe_float(data.get('natural_gas_therms')) or 0,
        'district_steam_mlb': _safe_float(data.get('district_steam_mlb')) or 0,
        'fuel_oil_2_gal':     _safe_float(data.get('fuel_oil_2_gal'))     or 0,
        'fuel_oil_4_gal':     _safe_float(data.get('fuel_oil_4_gal'))     or 0,
    }
    occupancy_groups = []
    for og in (data.get('occupancy_groups') or []):
        pt   = og.get('property_type', '').strip()
        area = _safe_float(og.get('floor_area'))
        if pt and area and area > 0:
            occupancy_groups.append({'property_type': pt, 'floor_area': area})
    if not occupancy_groups:
        return jsonify({'error': 'occupancy_groups required'}), 400

    current_placements_raw = data.get('current_placements') or []

    conn = get_saved_db_connection()
    measure_rows = conn.execute(
        'SELECT * FROM measures WHERE building_save_name = ? ORDER BY id', (building,)
    ).fetchall()
    conn.close()

    all_measures = [dict(m) for m in measure_rows]
    if not all_measures:
        return jsonify({'placements': []}), 200

    prices_cfg      = get_energy_prices_config()
    utility_factors = get_utility_factors(session.get('utility_factors', {}))

    if scope == 'unplaced':
        placed_ids         = {p.get('measure_id') for p in current_placements_raw if p.get('measure_id')}
        candidate_measures = [m for m in all_measures if m['id'] not in placed_ids]
    else:
        candidate_measures = all_measures

    if not candidate_measures:
        return jsonify({'placements': []}), 200

    if mode == 'minimize_fines':
        pre_cumulative = {k: 0 for k in ('electricity_kwh', 'natural_gas_therms',
                                          'district_steam_mlb', 'fuel_oil_2_gal', 'fuel_oil_4_gal')}
        if scope == 'unplaced' and current_placements_raw:
            m_map = {m['id']: m for m in all_measures}
            for p in current_placements_raw:
                m = m_map.get(p.get('measure_id'))
                if m:
                    pre_cumulative['electricity_kwh']    += m.get('elec_savings')  or 0
                    pre_cumulative['natural_gas_therms'] += m.get('gas_savings')   or 0
                    pre_cumulative['district_steam_mlb'] += m.get('steam_savings') or 0
                    pre_cumulative['fuel_oil_2_gal']     += m.get('oil2_savings')  or 0
                    pre_cumulative['fuel_oil_4_gal']     += m.get('oil4_savings')  or 0
        suggested = _suggest_minimize_fines(candidate_measures, energy, occupancy_groups,
                                            utility_factors, pre_cumulative)
    else:
        suggested = _suggest_minimize_cost(candidate_measures, energy, occupancy_groups,
                                           utility_factors, prices_cfg, use_discount, discount_rate)

    return jsonify({'placements': suggested})


def _suggest_minimize_fines(measures, energy, occupancy_groups, utility_factors, pre_cumulative=None):
    """Period-by-period greedy: select highest-GHG-saving measures per non-compliant period."""
    PERIOD_FIRST = [('2024_2029', 2024), ('2030_2034', 2030), ('2035_2039', 2035),
                    ('2040_2049', 2040), ('2050_plus', 2050)]
    uc = UNIT_CONVERSIONS

    cumulative = dict(pre_cumulative) if pre_cumulative else {
        k: 0 for k in ('electricity_kwh', 'natural_gas_therms',
                        'district_steam_mlb', 'fuel_oil_2_gal', 'fuel_oil_4_gal')
    }
    placed_ids = set()
    result     = []

    for period, first_year in PERIOD_FIRST:
        adj = {
            'electricity_kwh':    max(0, energy['electricity_kwh']    - cumulative['electricity_kwh']),
            'natural_gas_therms': max(0, energy['natural_gas_therms'] - cumulative['natural_gas_therms']),
            'district_steam_mlb': max(0, energy['district_steam_mlb'] - cumulative['district_steam_mlb']),
            'fuel_oil_2_gal':     max(0, energy['fuel_oil_2_gal']     - cumulative['fuel_oil_2_gal']),
            'fuel_oil_4_gal':     max(0, energy['fuel_oil_4_gal']     - cumulative['fuel_oil_4_gal']),
        }
        emissions = calculate_emissions(adj, utility_factors)[period]['total']
        limit     = calculate_limit(occupancy_groups, period)
        if emissions <= limit:
            continue

        deficit = emissions - limit
        f = utility_factors[period]

        def _ghg(m, _f=f):
            return (
                (m.get('elec_savings')  or 0) * _f['electricity_kwh'] +
                (m.get('gas_savings')   or 0) * uc['natural_gas_therm_to_kbtu']  * _f['natural_gas_kbtu'] +
                (m.get('steam_savings') or 0) * uc['district_steam_mlb_to_kbtu'] * _f['district_steam_kbtu'] +
                (m.get('oil2_savings')  or 0) * uc['fuel_oil_2_gal_to_kbtu']     * _f['fuel_oil_2_kbtu'] +
                (m.get('oil4_savings')  or 0) * uc['fuel_oil_4_gal_to_kbtu']     * _f['fuel_oil_4_kbtu']
            )

        candidates = sorted(
            [(m, _ghg(m)) for m in measures if m['id'] not in placed_ids],
            key=lambda x: x[1], reverse=True
        )
        for m, ghg_sav in candidates:
            if deficit <= 0:
                break
            result.append({'measure_id': m['id'], 'year': first_year})
            placed_ids.add(m['id'])
            cumulative['electricity_kwh']    += m.get('elec_savings')  or 0
            cumulative['natural_gas_therms'] += m.get('gas_savings')   or 0
            cumulative['district_steam_mlb'] += m.get('steam_savings') or 0
            cumulative['fuel_oil_2_gal']     += m.get('oil2_savings')  or 0
            cumulative['fuel_oil_4_gal']     += m.get('oil4_savings')  or 0
            deficit -= ghg_sav

    return result


def _suggest_minimize_cost(measures, energy, occupancy_groups, utility_factors,
                            prices_cfg, use_discount, discount_rate):
    """Per-measure: find the latest year where cumulative benefit >= capex."""
    YEARS = list(range(2024, 2051))
    uc    = UNIT_CONVERSIONS

    baseline_fines = {}
    for year in YEARS:
        period    = _get_period_for_year(year)
        emissions = calculate_emissions(energy, utility_factors)[period]['total']
        limit     = calculate_limit(occupancy_groups, period)
        baseline_fines[year] = calculate_penalty(emissions, limit)

    result = []

    for m in measures:
        capex = float(m.get('cost') or 0)

        # Free measures: place ASAP — no capital cost to optimise
        if capex <= 0:
            result.append({'measure_id': m['id'], 'year': 2024})
            continue

        ghg_cache = {}
        def _ghg_period(period, _m=m):
            if period not in ghg_cache:
                f = utility_factors[period]
                ghg_cache[period] = (
                    (_m.get('elec_savings')  or 0) * f['electricity_kwh'] +
                    (_m.get('gas_savings')   or 0) * uc['natural_gas_therm_to_kbtu']  * f['natural_gas_kbtu'] +
                    (_m.get('steam_savings') or 0) * uc['district_steam_mlb_to_kbtu'] * f['district_steam_kbtu'] +
                    (_m.get('oil2_savings')  or 0) * uc['fuel_oil_2_gal_to_kbtu']     * f['fuel_oil_2_kbtu'] +
                    (_m.get('oil4_savings')  or 0) * uc['fuel_oil_4_gal_to_kbtu']     * f['fuel_oil_4_kbtu']
                )
            return ghg_cache[period]

        def _annual_benefit(year, _m=m):
            period = _get_period_for_year(year)
            yp = {k: get_escalated_price(v['price'], v['escalator'], year) for k, v in prices_cfg.items()}
            e_sav = (
                (_m.get('elec_savings')  or 0) * yp.get('electricity_kwh', 0) +
                (_m.get('gas_savings')   or 0) * yp.get('natural_gas_therm', 0) +
                (_m.get('steam_savings') or 0) * yp.get('district_steam_mlb', 0) +
                (_m.get('oil2_savings')  or 0) * yp.get('fuel_oil_2_gal', 0) +
                (_m.get('oil4_savings')  or 0) * yp.get('fuel_oil_4_gal', 0)
            )
            fine_sav = min(_ghg_period(period) * PENALTY_RATE, baseline_fines.get(year, 0))
            return e_sav + fine_sav

        # Iterate 2050→2024; first match when iterating backward = latest breakeven year
        best_year = None
        for yr in reversed(YEARS):
            total = sum(
                _annual_benefit(t) / ((1 + discount_rate) ** (t - yr))
                if (use_discount and discount_rate > 0) else _annual_benefit(t)
                for t in range(yr, 2051)
            )
            if total >= capex:
                best_year = yr
                break

        if best_year is not None:
            result.append({'measure_id': m['id'], 'year': best_year})

    return result


# ── Measures upload & template ───────────────────────────────────────────────

@app.route('/api/upload-measures', methods=['POST'])
def upload_measures():
    building = request.form.get('building_save_name', '').strip()
    if not building:
        return jsonify({'error': 'building_save_name required'}), 400
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    filename = (f.filename or '').lower()
    rows_raw = []

    try:
        if filename.endswith('.csv'):
            import csv, io
            content = f.read().decode('utf-8-sig')
            reader  = csv.DictReader(io.StringIO(content))
            rows_raw = list(reader)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows_raw.append(dict(zip(headers, row)))
        else:
            return jsonify({'error': 'Only .csv or .xlsx files accepted'}), 400
    except Exception as exc:
        return jsonify({'error': f'File parsing error: {exc}'}), 400

    # Flexible header map (lower-case key → field name)
    HMAP = {
        'measure name':                    'name',
        'name':                            'name',
        'cost ($)':                        'cost',
        'cost':                            'cost',
        'electricity savings (kwh)':       'elec_savings',
        'electricity savings':             'elec_savings',
        'elec savings (kwh)':              'elec_savings',
        'natural gas savings (therms)':    'gas_savings',
        'natural gas savings':             'gas_savings',
        'gas savings (therms)':            'gas_savings',
        'steam savings (mlbs)':            'steam_savings',
        'steam savings (mlb)':             'steam_savings',
        'steam savings':                   'steam_savings',
        'oil #2 savings (gal)':            'oil2_savings',
        'oil #2 savings':                  'oil2_savings',
        'oil2 savings (gal)':              'oil2_savings',
        'oil #4 savings (gal)':            'oil4_savings',
        'oil #4 savings':                  'oil4_savings',
        'oil4 savings (gal)':              'oil4_savings',
    }

    conn    = get_saved_db_connection()
    create_saved_tables(conn)
    created = []
    errors  = []

    for i, row in enumerate(rows_raw, start=2):
        norm = {}
        for k, v in row.items():
            field = HMAP.get(str(k or '').strip().lower())
            if field:
                norm[field] = v
        name = str(norm.get('name') or '').strip()
        if not name:
            errors.append(f'Row {i}: missing Measure Name — skipped')
            continue
        cur = conn.execute('''
            INSERT INTO measures
                (building_save_name, name, cost, elec_savings, gas_savings,
                 steam_savings, oil2_savings, oil4_savings)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            building, name,
            _safe_float(norm.get('cost'))          or 0,
            _safe_float(norm.get('elec_savings'))  or 0,
            _safe_float(norm.get('gas_savings'))   or 0,
            _safe_float(norm.get('steam_savings')) or 0,
            _safe_float(norm.get('oil2_savings'))  or 0,
            _safe_float(norm.get('oil4_savings'))  or 0,
        ))
        created.append(dict(conn.execute(
            'SELECT * FROM measures WHERE id = ?', (cur.lastrowid,)
        ).fetchone()))

    conn.commit()
    conn.close()
    return jsonify({'created': created, 'errors': errors})


@app.route('/api/measures/template')
def measures_template():
    import csv, io
    from flask import Response
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow([
        'Measure Name', 'Cost ($)',
        'Electricity Savings (kWh)', 'Natural Gas Savings (therms)',
        'Steam Savings (mLbs)', 'Oil #2 Savings (gal)', 'Oil #4 Savings (gal)',
    ])
    w.writerow(['LED Lighting Retrofit',  50000, 100000, 0,    0, 0, 0])
    w.writerow(['Boiler Replacement',    120000,      0, 5000, 0, 0, 0])
    w.writerow(['Solar PV Installation', 200000,  80000, 0,    0, 0, 0])
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=reduction_measures_template.csv'},
    )


@app.route('/api/db-status')
def db_status():
    if not os.path.exists(DB_PATH):
        return jsonify({'initialized': False, 'count': 0, 'needs_reimport': False})
    conn = get_db_connection()
    count = conn.execute('SELECT COUNT(*) FROM buildings').fetchone()[0]
    # Check whether per-use floor area data was imported (indicates fresh import)
    has_floor_areas = False
    if count > 0:
        n = conn.execute(
            'SELECT COUNT(*) FROM buildings WHERE primary_floor_area IS NOT NULL AND primary_floor_area > 0'
        ).fetchone()[0]
        has_floor_areas = n > 0
    conn.close()
    return jsonify({
        'initialized':    count > 0,
        'count':          count,
        'needs_reimport': count > 0 and not has_floor_areas,
    })


def _safe_float(val):
    try:
        return float(val) if val not in (None, '') else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# CLI: initialize database
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--init-db',  action='store_true',
                        help='Download and import LL84 data (skips if DB already populated)')
    parser.add_argument('--reimport', action='store_true',
                        help='Drop and re-import LL84 data from scratch (picks up new columns)')
    parser.add_argument('--port', type=int, default=5000)
    args = parser.parse_args()

    if args.reimport:
        print('Re-importing LL84 database (this will replace existing data)...')
        import_ll84_data(verbose=True, force=True)
    elif args.init_db:
        print('Initializing LL84 database...')
        import_ll84_data(verbose=True)
    else:
        print(f'Starting LL97 Calculator on http://0.0.0.0:{args.port}')
        print('Tip: run with --init-db to download building database, or --reimport to refresh it.')
        app.run(host='0.0.0.0', port=args.port, debug=True)
