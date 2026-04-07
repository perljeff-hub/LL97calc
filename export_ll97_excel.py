"""
LL97 Compliance Excel Export
-----------------------------
Reads ll84.db, filters by year (default 2024), and produces an .xlsx file
where ALL emissions, limit, overage, and penalty calculations are live
Excel formulas referencing a Factors lookup sheet — no hardcoded calc values.

Usage:
    python export_ll97_excel.py [--year 2024] [--limit 25] [--all]
    python export_ll97_excel.py --all             # full LL84 dataset
"""

import sys, os, sqlite3, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from data.emission_factors import ESPM_LIMITS, DEFAULT_UTILITY_FACTORS, PENALTY_RATE

DB_PATH  = os.path.join(os.path.dirname(__file__), 'll84.db')
PERIOD   = '2024_2029'
FACTORS  = DEFAULT_UTILITY_FACTORS[PERIOD]

# ── Colour palette ────────────────────────────────────────────────────────
DARK_BLUE  = '1F3864'
MED_BLUE   = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
ORANGE     = 'F4B942'
RED_FILL   = 'FDECEA'
GREEN_FILL = 'E9F7EF'
WHITE      = 'FFFFFF'

def hex_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def bold_font(color=WHITE, size=10):
    return Font(bold=True, color=color, size=size)

def thin_border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────
# SHEET 1 — Factors  (lookup tables that formulas reference)
# ─────────────────────────────────────────────────────────────────────────
#
#  Row  Col A                          Col B
#  1    "LL97 Emission Factors ..."    (title)
#  2    "Compliance Period:"           2024-2029
#  3    [blank]
#  4    — Utility GHG Coefficients —
#  5    Label                          Factor
#  6    Electricity (kWh)              0.000288962   ← Factors!$B$6
#  7    Natural Gas (kBtu)             0.00005311    ← Factors!$B$7
#  8    District Steam (kBtu)          0.00004493    ← Factors!$B$8
#  9    Fuel Oil #2 (kBtu)             0.00007421    ← Factors!$B$9
#  10   Fuel Oil #4 (kBtu)             0.00007529    ← Factors!$B$10
#  11   [blank]
#  12   Penalty Rate ($/tCO2e)         268           ← Factors!$B$12
#  13   [blank]
#  14   — ESPM 2024-2029 Limits (tCO2e/sf/yr) —
#  15   Property Type                  Limit Factor
#  16+  <each type>                    <factor>      ← VLOOKUP range

UTIL_FACTOR_ROWS = {          # row number → (label, value)
    6:  ('Electricity (kWh)',       FACTORS['electricity_kwh']),
    7:  ('Natural Gas (kBtu)',      FACTORS['natural_gas_kbtu']),
    8:  ('District Steam (kBtu)',   FACTORS['district_steam_kbtu']),
    9:  ('Fuel Oil #2 (kBtu)',      FACTORS['fuel_oil_2_kbtu']),
    10: ('Fuel Oil #4 (kBtu)',      FACTORS['fuel_oil_4_kbtu']),
}
PENALTY_ROW   = 12
ESPM_HDR_ROW  = 14
ESPM_COL_HDR  = 15
ESPM_DATA_ROW = 16          # first property-type row


def build_factors_sheet(wb):
    ws = wb.create_sheet('Factors')
    ws.sheet_view.showGridLines = False

    # Title
    ws['A1'] = 'LL97 Emission Factors & ESPM Limits'
    ws['A1'].font = Font(bold=True, color=DARK_BLUE, size=13)
    ws['A2'] = 'Compliance Period:'
    ws['B2'] = '2024–2029'
    ws['B2'].font = Font(bold=True, color=MED_BLUE, size=11)

    # Utility factors header
    ws['A4'] = 'Utility GHG Emission Coefficients'
    ws['A4'].font = Font(bold=True, color=WHITE, size=10)
    ws['A4'].fill = hex_fill(MED_BLUE)
    ws['B4'] = 'tCO2e per unit'
    ws['B4'].font = Font(bold=True, color=WHITE, size=10)
    ws['B4'].fill = hex_fill(MED_BLUE)

    ws['A5'] = 'Energy Source'
    ws['B5'] = 'Factor'
    for cell in (ws['A5'], ws['B5']):
        cell.font  = Font(bold=True, size=9)
        cell.fill  = hex_fill(LIGHT_BLUE)
        cell.border = thin_border()

    for row, (label, value) in UTIL_FACTOR_ROWS.items():
        ws.cell(row, 1, label).border  = thin_border()
        c = ws.cell(row, 2, value)
        c.number_format = '0.00000000'
        c.border = thin_border()

    # Penalty rate
    ws['A11'] = ''
    ws['A12'] = 'Penalty Rate ($/tCO2e over limit)'
    ws['A12'].font = Font(bold=True, color=WHITE)
    ws['A12'].fill = hex_fill(MED_BLUE)
    ws.cell(PENALTY_ROW, 2, PENALTY_RATE).fill   = hex_fill(MED_BLUE)
    ws.cell(PENALTY_ROW, 2).font  = Font(bold=True, color=WHITE)
    ws.cell(PENALTY_ROW, 2).number_format = '$#,##0'
    ws['A12'].border = thin_border()
    ws.cell(PENALTY_ROW, 2).border = thin_border()

    # ESPM limits header
    ws['A14'] = 'ESPM Property Type Emission Intensity Limits — 2024–2029'
    ws['A14'].font = Font(bold=True, color=WHITE)
    ws['A14'].fill = hex_fill(MED_BLUE)
    ws['B14'].fill = hex_fill(MED_BLUE)
    ws['A15'] = 'Property Type'
    ws['B15'] = 'tCO2e / sf / yr  (2024–2029)'
    for cell in (ws['A15'], ws['B15']):
        cell.font  = Font(bold=True, size=9)
        cell.fill  = hex_fill(LIGHT_BLUE)
        cell.border = thin_border()

    espm_sorted = sorted(ESPM_LIMITS.items())
    for i, (ptype, periods) in enumerate(espm_sorted):
        r = ESPM_DATA_ROW + i
        ws.cell(r, 1, ptype).border = thin_border()
        c = ws.cell(r, 2, periods.get(PERIOD, 0))
        c.number_format = '0.00000'
        c.border = thin_border()

    espm_last_row = ESPM_DATA_ROW + len(espm_sorted) - 1

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 30

    # Return the range string for VLOOKUP (absolute, sheet-qualified)
    return f"Factors!$A${ESPM_DATA_ROW}:$B${espm_last_row}"


# ─────────────────────────────────────────────────────────────────────────
# SHEET 2 — Buildings
# ─────────────────────────────────────────────────────────────────────────
#
# Raw-data columns  (A–P):
#   A  BBL
#   B  Address
#   C  Borough
#   D  Gross Floor Area (sf)
#   E  Primary Property Type
#   F  Primary Floor Area (sf)   — blank → uses GFA
#   G  Secondary Property Type
#   H  Secondary Floor Area (sf)
#   I  Third Property Type
#   J  Third Floor Area (sf)
#   K  Electricity (kWh)
#   L  Natural Gas (kBtu)
#   M  District Steam (kBtu)
#   N  Fuel Oil #2 (kBtu)
#   O  Fuel Oil #4 (kBtu)
#   P  Reported GHG (tCO2e — LL84 self-reported, for reference only)
#
# Calculated columns  (Q–X) — ALL Excel formulas:
#   Q  Emissions – Electricity  =IF(K2="",0,K2*Factors!$B$6)
#   R  Emissions – Natural Gas  =IF(L2="",0,L2*Factors!$B$7)
#   S  Emissions – Dist Steam   =IF(M2="",0,M2*Factors!$B$8)
#   T  Emissions – Fuel Oil #2  =IF(N2="",0,N2*Factors!$B$9)
#   U  Emissions – Fuel Oil #4  =IF(O2="",0,O2*Factors!$B$10)
#   V  Total Emissions (tCO2e)  =SUM(Q2:U2)
#   W  Primary Limit contrib    =IF(E2="",0, IFERROR(VLOOKUP(E2,<espm_rng>,2,0),0) * IF(F2="",D2,F2))
#   X  Secondary Limit contrib  =IF(G2="",0, IFERROR(VLOOKUP(G2,<espm_rng>,2,0),0) * H2)
#   Y  Third Limit contrib      =IF(I2="",0, IFERROR(VLOOKUP(I2,<espm_rng>,2,0),0) * J2)
#   Z  Total LL97 Limit (tCO2e) =W2+X2+Y2
#   AA Overage (tCO2e)          =MAX(0, V2-Z2)
#   AB Annual Penalty ($)       =AA2 * Factors!$B$12

RAW_COLS = [
    ('BBL',                       'bbl'),
    ('Address',                   'address'),
    ('Borough',                   'borough'),
    ('Gross Floor Area (sf)',      'gross_floor_area'),
    ('Primary Property Type',     'primary_property_type'),
    ('Primary Floor Area (sf)',   'primary_floor_area'),
    ('Secondary Property Type',   'second_property_type'),
    ('Secondary Floor Area (sf)', 'second_floor_area'),
    ('Third Property Type',       'third_property_type'),
    ('Third Floor Area (sf)',     'third_floor_area'),
    ('Electricity (kWh)',         'electricity_kwh'),
    ('Natural Gas (kBtu)',        'natural_gas_kbtu'),
    ('District Steam (kBtu)',     'district_steam_kbtu'),
    ('Fuel Oil #2 (kBtu)',        'fuel_oil_2_kbtu'),
    ('Fuel Oil #4 (kBtu)',        'fuel_oil_4_kbtu'),
    ('Reported GHG tCO2e (ref.)', 'reported_ghg_emissions'),
]

CALC_HEADERS = [
    'Emiss – Electricity (tCO2e)',
    'Emiss – Natural Gas (tCO2e)',
    'Emiss – Dist Steam (tCO2e)',
    'Emiss – Fuel Oil #2 (tCO2e)',
    'Emiss – Fuel Oil #4 (tCO2e)',
    'Total Emissions (tCO2e)',
    'Limit – Primary (tCO2e)',
    'Limit – Secondary (tCO2e)',
    'Limit – Third (tCO2e)',
    'Total LL97 Limit (tCO2e)',
    'Overage (tCO2e)',
    'Annual Penalty ($)',
]

# Column index mapping (1-based)
C = {h: i+1 for i, (h, _) in enumerate(RAW_COLS)}   # raw data cols
# Calculated columns start right after raw cols
CALC_START = len(RAW_COLS) + 1   # = 17 → column Q

def calc_col(n):
    """Return column letter for the n-th calculated column (0-based)."""
    return get_column_letter(CALC_START + n)

# Convenience col letters for formulas
cBBL    = get_column_letter(C['BBL'])
cGFA    = get_column_letter(C['Gross Floor Area (sf)'])
cPT1    = get_column_letter(C['Primary Property Type'])
cA1     = get_column_letter(C['Primary Floor Area (sf)'])
cPT2    = get_column_letter(C['Secondary Property Type'])
cA2     = get_column_letter(C['Secondary Floor Area (sf)'])
cPT3    = get_column_letter(C['Third Property Type'])
cA3     = get_column_letter(C['Third Floor Area (sf)'])
cELEC   = get_column_letter(C['Electricity (kWh)'])
cNG     = get_column_letter(C['Natural Gas (kBtu)'])
cSTEAM  = get_column_letter(C['District Steam (kBtu)'])
cFO2    = get_column_letter(C['Fuel Oil #2 (kBtu)'])
cFO4    = get_column_letter(C['Fuel Oil #4 (kBtu)'])


def row_formulas(r, espm_rng):
    """Return list of formula strings for calculated columns at row r."""
    F = 'Factors'   # sheet name alias
    f6, f7, f8, f9, f10 = (
        f'{F}!$B$6', f'{F}!$B$7', f'{F}!$B$8',
        f'{F}!$B$9', f'{F}!$B$10',
    )
    pen = f'{F}!$B$12'

    # Emissions per fuel
    q = f'=IF({cELEC}{r}="",0,{cELEC}{r}*{f6})'
    rr= f'=IF({cNG}{r}="",0,{cNG}{r}*{f7})'
    s = f'=IF({cSTEAM}{r}="",0,{cSTEAM}{r}*{f8})'
    t = f'=IF({cFO2}{r}="",0,{cFO2}{r}*{f9})'
    u = f'=IF({cFO4}{r}="",0,{cFO4}{r}*{f10})'

    qc, rc, sc, tc, uc = [calc_col(i) for i in range(5)]

    # Total emissions
    v = f'=SUM({qc}{r}:{uc}{r})'
    vc = calc_col(5)

    # Per-use-type limit contributions
    # Primary: if no primary area reported, fall back to GFA
    w = (f'=IF({cPT1}{r}="",0,'
         f'IFERROR(VLOOKUP({cPT1}{r},{espm_rng},2,FALSE),0)'
         f'*IF(OR({cA1}{r}="",{cA1}{r}=0),{cGFA}{r},{cA1}{r}))')
    x = (f'=IF({cPT2}{r}="",0,'
         f'IFERROR(VLOOKUP({cPT2}{r},{espm_rng},2,FALSE),0)'
         f'*IF(OR({cA2}{r}="",{cA2}{r}=0),0,{cA2}{r}))')
    y = (f'=IF({cPT3}{r}="",0,'
         f'IFERROR(VLOOKUP({cPT3}{r},{espm_rng},2,FALSE),0)'
         f'*IF(OR({cA3}{r}="",{cA3}{r}=0),0,{cA3}{r}))')

    wc, xc, yc = calc_col(6), calc_col(7), calc_col(8)

    # Total limit
    z = f'={wc}{r}+{xc}{r}+{yc}{r}'
    zc = calc_col(9)

    # Overage
    aa = f'=IF({zc}{r}=0,"",MAX(0,{vc}{r}-{zc}{r}))'
    aac = calc_col(10)

    # Penalty
    ab = f'=IF({zc}{r}=0,"",{aac}{r}*{pen})'

    return [q, rr, s, t, u, v, w, x, y, z, aa, ab]


def build_buildings_sheet(wb, rows, espm_rng, preview_n=None):
    ws = wb.active
    ws.title = 'Buildings'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    total_raw  = len(RAW_COLS)
    total_calc = len(CALC_HEADERS)
    total_cols = total_raw + total_calc

    # ── Header row ──────────────────────────────────────────────────────
    for ci, (label, _) in enumerate(RAW_COLS, start=1):
        cell = ws.cell(1, ci, label)
        cell.font  = bold_font(WHITE)
        cell.fill  = hex_fill(DARK_BLUE)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.border = thin_border()

    for ci, label in enumerate(CALC_HEADERS, start=CALC_START):
        cell = ws.cell(1, ci, label)
        cell.font  = bold_font(WHITE)
        cell.fill  = hex_fill(MED_BLUE)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.border = thin_border()

    ws.row_dimensions[1].height = 40

    # ── Data rows ────────────────────────────────────────────────────────
    display_rows = rows if preview_n is None else rows[:preview_n]

    for ri, row in enumerate(display_rows, start=2):
        is_even = ri % 2 == 0

        # Raw data
        for ci, (_, field) in enumerate(RAW_COLS, start=1):
            val = row[field]
            cell = ws.cell(ri, ci, val)
            cell.border = thin_border()
            cell.fill = hex_fill('F5F9FF' if is_even else WHITE)
            if field in ('gross_floor_area', 'primary_floor_area',
                         'second_floor_area', 'third_floor_area',
                         'electricity_kwh', 'natural_gas_kbtu',
                         'district_steam_kbtu', 'fuel_oil_2_kbtu',
                         'fuel_oil_4_kbtu', 'reported_ghg_emissions'):
                cell.number_format = '#,##0.0'

        # Calculated formulas
        formulas = row_formulas(ri, espm_rng)
        for fi, formula in enumerate(formulas):
            ci = CALC_START + fi
            cell = ws.cell(ri, ci, formula)
            cell.border = thin_border()

            # Formatting per column
            header = CALC_HEADERS[fi]
            if 'Penalty' in header:
                cell.number_format = '$#,##0'
                # Conditional-style: positive penalty → orange fill
                # (openpyxl static; for true conditional formatting see below)
            elif 'tCO2e' in header:
                cell.number_format = '#,##0.00'

            cell.fill = hex_fill('FFF3E0' if 'Penalty' in header else
                                 ('F5F9FF' if is_even else WHITE))

    # ── Column widths ────────────────────────────────────────────────────
    widths = {
        cBBL:   14,
        get_column_letter(C['Address']): 32,
        get_column_letter(C['Borough']): 10,
        cGFA:   14,
        cPT1:   28, cA1: 14,
        cPT2:   28, cA2: 14,
        cPT3:   28, cA3: 14,
        cELEC:  14, cNG: 14, cSTEAM: 16, cFO2: 16, cFO4: 16,
        get_column_letter(C['Reported GHG tCO2e (ref.)']): 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    # Calculated columns
    for i in range(total_calc):
        ws.column_dimensions[calc_col(i)].width = 18

    # ── Conditional formatting: penalty column red if > 0 ────────────────
    from openpyxl.formatting.rule import CellIsRule
    pen_col = calc_col(11)   # Annual Penalty column
    over_col = calc_col(10)  # Overage column
    last_data_row = len(display_rows) + 1

    red_f  = PatternFill('solid', fgColor='FDDCDC')
    green_f= PatternFill('solid', fgColor='DFF0D8')
    ws.conditional_formatting.add(
        f'{pen_col}2:{pen_col}{last_data_row}',
        CellIsRule(operator='greaterThan', formula=['0'], fill=red_f)
    )
    ws.conditional_formatting.add(
        f'{over_col}2:{over_col}{last_data_row}',
        CellIsRule(operator='greaterThan', formula=['0'], fill=red_f)
    )
    # Green if compliant (overage = 0)
    ws.conditional_formatting.add(
        f'{over_col}2:{over_col}{last_data_row}',
        CellIsRule(operator='equal', formula=['0'], fill=green_f)
    )

    return ws


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--year',  default='2024')
    parser.add_argument('--limit', type=int, default=25,
                        help='Rows to export (default 25; 0 = all)')
    parser.add_argument('--all',   action='store_true', help='Export full dataset')
    parser.add_argument('--out',   default='', help='Output .xlsx filename')
    args = parser.parse_args()

    if not os.path.exists(DB_PATH) or os.path.getsize(DB_PATH) == 0:
        print(f'ERROR: {DB_PATH} is missing or empty.')
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    rows = conn.execute('''
        SELECT * FROM buildings
        WHERE year_ending LIKE ?
        ORDER BY gross_floor_area DESC
    ''', (f'{args.year}%',)).fetchall()
    conn.close()

    total = len(rows)
    print(f'Found {total:,} buildings for year {args.year}')

    preview_n = None if (args.all or args.limit == 0) else args.limit
    export_n  = total if preview_n is None else min(preview_n, total)

    out_name = args.out or (
        f'll97_ll84_{args.year}_{"full" if preview_n is None else f"top{export_n}"}.xlsx'
    )
    out_path = os.path.join(os.path.dirname(__file__), out_name)

    wb = openpyxl.Workbook()
    espm_rng = build_factors_sheet(wb)
    build_buildings_sheet(wb, rows, espm_rng, preview_n=preview_n)

    # Move Buildings sheet to front
    wb.move_sheet('Buildings', offset=-wb.sheetnames.index('Buildings'))

    wb.save(out_path)
    print(f'Saved: {out_path}  ({export_n:,} buildings)')
    print(f'Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    main()
